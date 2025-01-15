import openpyxl
import yaml

from paths_manager import add_school_data_yaml


def read_excel(filepath, sheet_name):
    """
    读取excel数据
    :param filepath:
    :param sheet_name:
    :return:
    """
    # 获取整个文档对象
    wb = openpyxl.load_workbook(filepath)
    print(wb)
    # 获取某个sheet工作表的数据
    sheet_data = wb[sheet_name]
    print(sheet_data)
    # 获取某个单元格数据
    # print(sheet_data.cell(2, 2).value)
    lines_count = sheet_data.max_row  # 获取总行数
    cols_count = sheet_data.max_column  # 获取总列数
    # print(lines_count,cols_count)
    # 注意：openpyxl里读取时行号和列号都从是1开始
    data = []  # 用来存储所有行的数据，每行数据都是这个列表的子列表
    for l in range(2, lines_count + 1):  # l:2,3,4,5
        line = []  # 用来存储当前行所有列的单元格数据
        for c in range(1, cols_count + 1):  # c:1,2,3,4,5
            # print(sheet_data.cell(l,c).value)
            cell_data = sheet_data.cell(l, c).value
            line.append(cell_data)
        data.append(line)
    return data


def load_yaml_file(filepath):
    """
    读取yaml文件
    :param filepath:
    :return:
    """
    with open(file=filepath, mode='r', encoding='utf-8') as f:
        content = yaml.load(f, Loader=yaml.FullLoader)
        return content


def write_yaml(filepath, content):
    """

    :param filepath:
    :param content:
    :return:
    """
    with open(file=filepath, mode='w', encoding='utf-8') as f:
        yaml.dump(content, f, Dumper=yaml.Dumper)


# 这里调试是没有问题的，但是在run.py中是会报错的，路径错误
if __name__ == '__main__':
    print(read_excel('../data/add_school_data.xlsx', '添加学校测试数据'))  # 第一个参数是excel文件路径，第二个参数是sheet页名称
    print(load_yaml_file(add_school_data_yaml))
